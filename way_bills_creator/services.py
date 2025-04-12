import os
import calendar
from pyteamup import Calendar
from datetime import date, datetime
import pandas as pd
from config.settings import BASE_DIR, CALENDAR_ID, API_KEY
from way_bills_creator.models import Event, Car, Location, Driver, WayBill, Distance
from openpyxl import load_workbook
import logging

log_file = "logs_" + str(date.today()) + ".log"
logging.basicConfig(
    level=logging.WARNING,
    format="%(asctime)s - %(name)s - %(message)s",
    filename=log_file,
)

logger = logging.getLogger()

logger_wb_create = logging.getLogger("Создание путевых листов")
logger_dwnld_report = logging.getLogger("Выгрузка отчета в excel")
logger_dwnld_wbs_excl = logging.getLogger("Выгрузка путевых листов в excel")
logger_dwnld_events_TUp = logging.getLogger("Загрузка событий из TeamUp")
logger_dwnld_events_csv = logging.getLogger("Загрузка событий из .csv")
logger_dwnld_cars_csv = logging.getLogger("Загрузка авто из .csv")
logger_dwnld_drvrs_csv = logging.getLogger("Загрузка водителей из .csv")
logger_dwnld_locthns_csv = logging.getLogger("Загрузка мест из .csv")

period_list = (
    "январь",
    "февраль",
    "март",
    "I_квартал",
    "апрель",
    "май",
    "июнь",
    "II_квартал",
    "I_полугодие",
    "июль",
    "август",
    "сентябрь",
    "III_квартал",
    "9_месяцев",
    "октябрь",
    "ноябрь",
    "декабрь",
    "IV_квартал",
    "II_полугодие",
    "год",
)

year_list = (
    "2020",
    "2021",
    "2022",
    "2023",
    "2024",
    "2025",
    "2026",
    "2027",
    "2028",
    "2029",
    "2030",
    "2031",
    "2032",
)


class LogsService:
    """Класс для загрузки логов из файла"""

    @staticmethod
    def get_logs():
        """
        Загружает данные из log файла
        """
        file = os.path.join(BASE_DIR, log_file)
        if file:
            with open(file, "r", encoding="utf-8") as f:
                text = f.readlines()
            return text


class WaybillsService:
    """Класс сервиса Путевых листов"""

    @staticmethod
    def create_waybills(period, year):
        """Метод создания путевых листов для выбранного пользователем периода"""

        enddt = get_end_date(period, year)
        startdt = get_start_date(period, year)

        logger_wb_create.warning(
            f"Создаем путевые листы для периода с {startdt} по {enddt}"
        )
        daterange = pd.date_range(start=startdt, end=enddt)
        count = 0
        for dateselect in daterange:
            events = Event.objects.filter(
                start_date__lte=dateselect, end_date__gte=dateselect
            )
            for event in events:
                logger_wb_create.warning(f"Создаем путевой лист для События: {event}")
                logger_wb_create.warning(f"Ищем водителя {event.calendar_name}")
                try:
                    driver = Driver.objects.get(name=event.calendar_name)
                except Exception as ex:
                    logger_wb_create.warning(
                        f'{event.calendar_name}, " не водитель / ", {ex}'
                    )
                    continue
                car = Car.objects.get(id=driver.car.pk)
                location_from, location_to = get_location(event.location, event.subject)
                logger_wb_create.warning(
                    f"Получили Места из справочника {location_from}, {location_to}"
                )
                distance = get_distance(location_from, location_to)
                logger_wb_create.warning(f"Получили Расстояние - {distance}")
                logger_wb_create.warning("Проверяем есть ли уже такой путевой лист")

                waybill_exist = WayBill.objects.filter(
                    date=dateselect,
                    driver=driver,
                    car=car,
                    event_id=event.event_id,
                    location_from=location_from,
                    location_to=location_to,
                    distance=distance,
                )
                if waybill_exist:
                    logger_wb_create.warning(
                        f"Путевой лист уже существует - {waybill_exist}"
                    )
                else:
                    waybill = WayBill.objects.create(
                        date=dateselect,
                        driver=driver,
                        car=car,
                        event_id=event.event_id,
                        location_from=location_from,
                        location_to=location_to,
                        distance=distance,
                    )
                    waybill.save()
                    count += 1
                    logger_wb_create.warning(f"Создали новый путевой лист - {waybill}")
        logger_wb_create.warning(f"Создано {count} новых путевых листов")
        return count

    @staticmethod
    def download_report(selected_period_excel, selected_year_excel):
        """Метод создания и сохранения отчета по путевым листам и расходу топлива за
        выбранный пользователем период в Excel"""

        startdt = get_start_date(selected_period_excel, selected_year_excel)
        enddt = get_end_date(selected_period_excel, selected_year_excel)
        logger_dwnld_report.warning(
            f"Создаем отчеты для периода с {startdt} по {enddt}"
        )
        year = startdt.year
        count = 0
        for month in range(startdt.month, enddt.month + 1):
            cols = [
                "Дата ПЛ",
                "Номер ПЛ",
                "Водитель",
                "Номер авто",
                "Название авто",
                "СНИЛС",
                "Номер водительского удостоверения",
                "Дата выдачи вод. удост.",
                "Расход топлива на 100км",
                "Тип топлива",
                "Место от",
                "Название компании",
                "Адрес",
                "Место до",
                "Название компании",
                "Адрес",
                "Дистанция",
                "Итого дизельного топлива",
                "Итого автобензина",
            ]
            dat_full = pd.DataFrame(columns=cols)

            start_date = date(year, month, 1)
            end_date = date(year, month, calendar.monthrange(year, month)[1])

            waybills = WayBill.objects.filter(date__lte=end_date, date__gte=start_date)
            for waybill in waybills:
                if month in ["01", "02", "03", "11", "12"]:
                    l100km = waybill.car.l100km_w
                else:
                    l100km = waybill.car.l100km

                data = [
                    (
                        waybill.date,
                        waybill.pk,
                        waybill.driver.name,
                        waybill.car.car_num,
                        waybill.car.car_name,
                        waybill.driver.snils_num,
                        waybill.driver.lic_num,
                        waybill.driver.lic_date,
                        l100km,
                        waybill.car.oil_type,
                        waybill.location_from.location_name,
                        waybill.location_from.company_name,
                        waybill.location_from.address,
                        waybill.location_to.location_name,
                        waybill.location_to.company_name,
                        waybill.location_to.address,
                        waybill.distance,
                        0.0,
                        0.0,
                    )
                ]

                dataf = pd.DataFrame(data, columns=cols)
                dataf.loc[dataf["Тип топлива"] == "D", ["Итого дизельного топлива"]] = (
                    waybill.distance * 2 * l100km / 100
                )
                dataf.loc[dataf["Тип топлива"] == "G", ["Итого автобензина"]] = (
                    waybill.distance * 2 * l100km / 100
                )
                dat_full = pd.concat([dat_full, dataf], ignore_index=True)
            if waybills:
                logger_dwnld_report.warning(
                    f"Сохраняем отчет в файл Reports/output_{start_date.strftime('%Y.%m')}.xlsx"
                )
                dat_full.to_excel(
                    "Reports/output_" + start_date.strftime("%Y.%m") + ".xlsx"
                )
                count += 1
        logger_dwnld_report.warning(f"Создано {count} отчетов")
        return count

    @staticmethod
    def download_wbs_excel(selected_period, selected_year):
        """Метод создания печатных форм путевых листов и сохранение их в Excel за выбранный пользователем период"""
        enddt = get_end_date(selected_period, selected_year)
        startdt = get_start_date(selected_period, selected_year)
        logger_dwnld_wbs_excl.warning(
            f"Выгружаем путевые листы в Excel для периода с {startdt} по {enddt}"
        )
        waybills = WayBill.objects.filter(date__lte=enddt, date__gte=startdt)
        count = 0
        for waybill in waybills:
            logger_dwnld_wbs_excl.warning(waybill)
            rumonth = [
                "",
                "январь",
                "февраль",
                "март",
                "апрель",
                "май",
                "июнь",
                "июль",
                "август",
                "сентябрь",
                "октябрь",
                "ноябрь",
                "декабрь",
            ]
            waybill_date = pd.to_datetime(waybill.date, format="%d.%m.%Y")

            path = (
                "Reports/"
                + str(waybill.car.car_num)
                + "/"
                + str(waybill_date.year)
                + "_"
                + rumonth[waybill_date.month]
            )

            if not os.path.exists(path):
                os.makedirs(path)

            waybill_date_str = waybill.date.strftime("%d.%m.%Y")
            filename = path + "/wb_" + waybill_date_str
            template = "template_wb"

            if not os.path.exists(filename):
                workbook = load_workbook(f"{template}.xlsx")
            else:
                workbook = load_workbook(filename + ".xlsx")

            sheet = workbook.active
            if waybill_date.strftime("%m") in ["01", "02", "03", "11", "12"]:
                l100km = waybill.car.l100km_w
            else:
                l100km = waybill.car.l100km

            sheet["A1"] = (
                f"Путевой лист легкового автомобиля № {waybill.id} от "
                + waybill_date_str
            )

            sheet["C5"] = (
                f"Легковой автомобиль {waybill.car.car_name}, "
                f"государственный регистрационный знак {waybill.car.car_num}"
            )
            sheet["C7"] = (
                f"{waybill.driver.name}, СНИЛС {waybill.driver.snils_num}, "
                f"водительское удостоверение № {waybill.driver.lic_num} от {waybill.driver.lic_date}"
            )

            sheet["B16"] = (
                f"{waybill.location_from.company_name}, {waybill.location_from.address}"
            )
            sheet["L18"] = (
                f"{waybill.location_from.company_name}, {waybill.location_from.address}"
            )
            sheet["L16"] = (
                f"{waybill.location_to.company_name}, {waybill.location_to.address}"
            )
            sheet["B18"] = (
                f"{waybill.location_to.company_name}, {waybill.location_to.address}"
            )
            sheet["U16"] = waybill.distance
            sheet["U18"] = waybill.distance

            sheet["A21"] = f"Пробег автомобиля: {waybill.distance * 2} км"

            if waybill.car.oil_type == "D":
                sheet["A22"] = (
                    f"Расход топлива: {round(float(l100km), 2)} л марки Дизель"
                )
            if waybill.car.oil_type == "G":
                sheet["A22"] = (
                    f"Расход топлива: {round(float(l100km), 2)} л марки Автобензины"
                )

            workbook.save(filename + ".xlsx")
            logger_dwnld_wbs_excl.warning(
                f"Сохранили путевой лист {waybill} в файл {filename}.xlsx"
            )
            workbook.close()
            count += 1

        logger_dwnld_wbs_excl.warning(f"Сохранили {count} путевых листов")

        return count


def get_location(event_loc, subj):
    """Метод получения Места назначения"""
    locations = Location.objects.all()
    office_loc = Location.objects.get(location_name="Офис")
    location_to = office_loc
    for location in locations:
        str_subj = subj
        str_loc = event_loc
        str_location = location.location_name
        if (
            str_location.lower() in str_subj.lower()
            or str_location.lower() in str_loc.lower()
        ):
            location_to = location
    location_from, location_to = office_loc, location_to
    return location_from, location_to


def get_distance(location_from, location_to):
    """Метод получения расстояния между двумя местами назначения"""
    try:
        distance = Distance.objects.get(loc_from=location_from, loc_to=location_to)
    except Exception as ex:
        logger.error(ex)
        return 0.0
    return distance.loc_km


class CalendarService:
    @staticmethod
    def download(period, year):
        calendar_teamup = Calendar(CALENDAR_ID, API_KEY)
        enddt = get_end_date(period, year)
        startdt = get_start_date(period, year)

        logger_dwnld_events_TUp.warning(
            f"Выгружаем События из TeamUp для периода с {startdt} по {enddt}"
        )
        try:
            events = calendar_teamup.get_event_collection(
                start_dt=startdt, end_dt=enddt, returnas="dataframe"
            )
            logger_dwnld_events_TUp.warning("Прочитали События из TeamUp")
        except Exception as ex:
            return ex

        try:
            os.makedirs(f"{BASE_DIR}/backup", exist_ok=True)

            events.to_csv(
                f"{BASE_DIR}/backup/event_collection_{startdt}_{enddt}__{date.today()}.csv"
            )
            logger_dwnld_events_TUp.warning(
                f"Сохранили События в файл {BASE_DIR}/backup/event_collection_{startdt}_{enddt}__{date.today()}.csv"
            )
        except Exception as ex:
            return ex

        try:
            subcals = pd.DataFrame(calendar_teamup.subcalendars)
            logger_dwnld_events_TUp.warning("Прочитали Календари из TeamUp")
            logger_dwnld_events_TUp.warning(
                f"Сохранили Календари в файл {BASE_DIR}/backup/subcals_{startdt}_{enddt}__{date.today()}.csv"
            )
            subcals.to_csv(
                f"{BASE_DIR}/backup/subcals_{startdt}_{enddt}__{date.today()}.csv"
            )
        except Exception as ex:
            return ex

        count = 0
        for i, row in events.iterrows():
            events_exist = Event.objects.filter(event_id=row["id"])
            if events_exist:
                logger_dwnld_events_TUp.warning(
                    f"Удаляем существующие События с id= {row['id']}"
                )
                for event in events_exist:
                    logger_dwnld_events_TUp.warning(event)
                    event.delete()

            for cal_id in row["subcalendar_ids"]:
                calendar_name_indx = subcals[subcals.id == cal_id].index
                calendar_name = subcals.loc[calendar_name_indx[0], "name"]
                start_dt = datetime.strptime(row["start_dt"][:10], "%Y-%m-%d").date()
                end_dt = datetime.strptime(row["end_dt"][:10], "%Y-%m-%d").date()

                event = Event.objects.create(
                    event_id=row["id"],
                    start_date=start_dt,
                    end_date=end_dt,
                    subject=row["title"],
                    calendar_name=calendar_name,
                    location=row["location"],
                    who=row["who"],
                )
                event.save()
                count += 1
                logger_dwnld_events_TUp.warning(f"Записали новое Событие в БД: {event}")
        logger_dwnld_events_TUp.warning(f"Записали {count} новых Событий в БД")
        return count


def get_end_date(period, year):
    """Метод получения последней даты периода"""
    if period in ["I_квартал", "март"]:
        month = 3
    elif period in ["II_квартал", "I_полугодие", "июнь"]:
        month = 6
    elif period in ["III_квартал", "9_месяцев", "сентябрь"]:
        month = 9
    elif period in ["декабрь", "IV_квартал", "II_полугодие", "год"]:
        month = 12
    elif period == "январь":
        month = 1
    elif period == "февраль":
        month = 2
    elif period == "апрель":
        month = 4
    elif period == "май":
        month = 5
    elif period == "июль":
        month = 7
    elif period == "август":
        month = 8
    elif period == "октябрь":
        month = 10
    elif period == "ноябрь":
        month = 11
    else:
        month = 1

    last_day = calendar.monthrange(int(year), month)
    end_date = date(year=int(year), month=month, day=last_day[1])
    return end_date


def get_start_date(period, year):
    """Метод получения первой даты периода"""
    if period in ["I_квартал", "январь", "I_полугодие", "9_месяцев", "год"]:
        month = 1
    elif period in ["II_квартал", "апрель"]:
        month = 4
    elif period in ["III_квартал", "июль", "II_полугодие"]:
        month = 7
    elif period in ["октябрь", "IV_квартал"]:
        month = 10
    elif period == "март":
        month = 3
    elif period == "февраль":
        month = 2
    elif period == "декабрь":
        month = 12
    elif period == "май":
        month = 5
    elif period == "июнь":
        month = 6
    elif period == "август":
        month = 8
    elif period == "сентябрь":
        month = 9
    elif period == "ноябрь":
        month = 11
    else:
        month = 1

    start_date = date(year=int(year), month=month, day=1)
    return start_date


class CSVService:
    """Класс сервиса загрузки данных через файлы формата CSV"""

    @staticmethod
    def cars_load(file):
        """Метод загрузки автомобилей"""
        logger_dwnld_cars_csv.warning(f"Загружаем автомобили из файла {file}")
        try:
            car_data = pd.read_csv(
                file, usecols=["CarNum", "CarName", "l100km", "l100km_w", "OilType"]
            )
        except Exception as ex:
            return ex
        count = 0
        for i, row in car_data.iterrows():
            car = Car.objects.create(
                car_name=row["CarName"],
                car_num=row["CarNum"],
                oil_type=row["OilType"],
                l100km=row["l100km"],
                l100km_w=row["l100km_w"],
            )
            car.save()
            count += 1
            logger_dwnld_cars_csv.warning(f"Автомобиль добавлен:  {car}")
        logger_dwnld_cars_csv.warning(f"Автомобилей добавлено:  {count}")
        return count

    @staticmethod
    def locations_load(file):
        """Метод загрузки Мест назначения"""
        logger_dwnld_locthns_csv.warning(f"Загружаем Места из файла {file}")
        try:
            loc_data = pd.read_csv(file, usecols=["Location", "CompanyName", "Address"])
        except Exception as ex:
            return ex
        count = 0
        for i, row in loc_data.iterrows():
            location = Location.objects.create(
                location_name=row["Location"],
                company_name=row["CompanyName"],
                address=row["Address"],
            )
            location.save()
            count += 1
            logger_dwnld_locthns_csv.warning(f"Место добавлено:  {location}")
        logger_dwnld_locthns_csv.warning(f"Мест добавлено:  {count}")
        return count

    @staticmethod
    def drivers_load(file):
        """Метод загрузки данных по Водителям"""
        logger_dwnld_drvrs_csv.warning(f"Загружаем Водителей из файла {file}")
        try:
            drivers_data = pd.read_csv(
                file, usecols=["DriverName", "CarNum", "SNILSnum", "LicNum", "licDate"]
            )
        except Exception as ex:
            return ex
        count = 0
        for i, row in drivers_data.iterrows():
            car = Car.objects.get(car_num=row["CarNum"])
            driver = Driver.objects.create(
                name=row["DriverName"],
                car=car,
                snils_num=row["SNILSnum"],
                lic_num=row["LicNum"],
                lic_date=row["licDate"],
            )
            driver.save()
            count += 1
            logger_dwnld_drvrs_csv.warning(f"Водитель добавлен:  {driver}")
        logger_dwnld_drvrs_csv.warning(f"Водителей добавлено:  {count}")
        return count

    @staticmethod
    def events_load(file):
        """Метод загрузки Событий календаря"""
        logger_dwnld_events_csv.warning(f"Загружаем События из файла {file}")
        try:
            events_data = pd.read_csv(
                file,
                usecols=[
                    "Id",
                    "Subject",
                    "Start Date",
                    "End Date",
                    "Calendar Name",
                    "Who",
                    "Location",
                ],
            )
        except Exception as ex:
            return ex
        count = 0
        for i, row in events_data.iterrows():
            events_exist = Event.objects.filter(event_id=row["Id"])
            if events_exist:
                logger_dwnld_events_csv.warning(
                    f"Удаляем существующие события с id= {row['Id']}"
                )
                for event in events_exist:
                    logger_dwnld_events_csv.warning(f"Событие удалено: {event} ")
                    event.delete()

            for calendar_name in row["Calendar Name"].split("|"):
                start_dt = datetime.strptime(row["Start Date"], "%d/%m/%Y").date()
                end_dt = datetime.strptime(row["End Date"], "%d/%m/%Y").date()

                event = Event.objects.create(
                    event_id=row["Id"],
                    start_date=start_dt,
                    end_date=end_dt,
                    subject=row["Subject"],
                    calendar_name=calendar_name,
                    location=row["Location"],
                    who=row["Who"],
                )
                event.save()
                count += 1
                logger_dwnld_events_csv.warning(f"Событие добавлено:  {event}")
        logger_dwnld_events_csv.warning(f"Событий добавлено: {count}")
        return count
